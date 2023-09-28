import os
import openpyxl
import sys
import scipy.stats as stats
from scipy.stats import t
import numpy as np
# Calculation
p=0.95

def read_values_from_file(filename):
    with open(filename, 'r') as file:
        values = [float(line.strip()) for line in file]
    return values

def calculate_mean(values):
    return sum(values) / len(values)

def calculate_standard_deviation(values, mean):
    variance = sum((x - mean) ** 2 for x in values) / len(values)
    return variance ** 0.5

def calculate_confidence_interval(values, mean, std_deviation,p):
    n = len(values)
    t_value = stats.t.ppf(1 - (1 - p) / 2, n-1)
    margin_of_error = t_value * std_deviation / (n ** 0.5)
    return mean - margin_of_error, mean + margin_of_error

def calculate_median(values):
    sorted_values = sorted(values)
    n = len(values)
    if n % 2 == 0:
        median = (sorted_values[n // 2 - 1] + sorted_values[n // 2]) / 2
    else:
        median = sorted_values[n // 2]
    return median

def write_to_excel(file_path, data):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    column_header = "data"
    sheet.cell(row=1, column=1).value = column_header
    # Записываем данные в первый столбец
    for i, row in enumerate(data, start=2):
        sheet.cell(row=i, column=1).value = row
    workbook.save(file_path)


def discard_outliers(significance_level=0.05):
    data = values
    mean = np.mean(data)
    std = np.std(data)
    n = len(data)
    critical_value = t.ppf(1 - significance_level / 2, n - 1)
    lower_bound = mean - critical_value * std / np.sqrt(n)
    upper_bound = mean + critical_value * std / np.sqrt(n)
    filtered_data = [x for x in data if lower_bound <= x <= upper_bound]

    return filtered_data
# GUI
from tkinter.filedialog import askopenfilename, asksaveasfile
import tkinter as tk
import pandas as pd
from pandastable import Table

def button1_func():
    global input_filename
    global values
    global name_without_extension
    global data
    input_filename = askopenfilename(title='Файл с данными')
    values = read_values_from_file(input_filename)
    name_without_extension = os.path.splitext(input_filename)[0]
    write_to_excel(str(name_without_extension) + '.xlsx', values)
    data = pd.read_excel(str(name_without_extension) + '.xlsx')
    # Создание таблицы из DataFrame
    table = Table(frame, dataframe=data)
    table.show()
    os.remove(str(name_without_extension) + '.xlsx')
    button2 = tk.Button(window, text="Рассчитать", command=button2_func)
    button2.pack()



def button2_func():
    global mean
    global std_deviation
    global confidence_interval
    global median
    global values
    mean = calculate_mean(values)
    std_deviation = calculate_standard_deviation(values, mean)
    confidence_interval = calculate_confidence_interval(values, mean, std_deviation,p)
    median = calculate_median(values)
    label = tk.Label(window)
    label.config(text=f"file path: {input_filename}\n"
                   +f"Mean: {mean}\n"
                   + f"Standard Deviation: {std_deviation}\n"
                   + f"Confidence Interval: {confidence_interval}\n"
                   + f"Median: {median}\n")
    label.pack()
    button3 = tk.Button(window, text="Сохранить", command=button3_func)
    button3.pack()


def button3_func():
    new_file = asksaveasfile(title="Сохранить файл", defaultextension=".txt",
                             filetypes=(("Текстовый файл", "*.txt"),))
    new_file.write(f"file path: {input_filename}\n"
                   + f"Mean: {mean}\n"
                   + f"Standard Deviation: {std_deviation}\n"
                   + f"Confidence Interval: {confidence_interval}\n"
                   + f"Median: {median}\n"
                   + f"data:{values}")
    new_file.close()

def button4_func():
    values1=discard_outliers()
    write_to_excel(str(name_without_extension) + '.xlsx', values1)
    data = pd.read_excel(str(name_without_extension) + '.xlsx')
    os.remove(str(name_without_extension) + '.xlsx')
    mean = calculate_mean(values1)
    std_deviation = calculate_standard_deviation(values1, mean)
    confidence_interval = calculate_confidence_interval(values1, mean, std_deviation, p)
    median = calculate_median(values1)
    label = tk.Label(window)
    label.config(text=f"file path: {input_filename}\n"
                      + f"Mean: {mean}\n"
                      + f"Standard Deviation: {std_deviation}\n"
                      + f"Confidence Interval: {confidence_interval}\n"
                      + f"Median: {median}\n")
    label.pack()
    new_file = asksaveasfile(title="Сохранить файл", defaultextension=".txt",
                             filetypes=(("Текстовый файл", "*.txt"),))
    new_file.write(f"file path: {input_filename}\n"
                   + f"Mean: {mean}\n"
                   + f"Standard Deviation: {std_deviation}\n"
                   + f"Confidence Interval: {confidence_interval}\n"
                   + f"Median: {median}\n"
                   + f"data:{values1}")
    new_file.close()


window = tk.Tk()

button1 = tk.Button(window, text="Файл", command=button1_func)
button1.pack()

button4 = tk.Button(window, text="Отбросить аномальные", command=button4_func)
button4.pack()

frame = tk.Frame(window)
frame.pack()

def new_close_function():
    sys.exit()
window.protocol("WM_DELETE_WINDOW", new_close_function)

# Запуск главного цикла приложения
window.mainloop()

